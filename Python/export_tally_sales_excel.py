import sys, json, os, time, traceback
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from db import get_connection


def autosize(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def main():
    start_clock = datetime.now().strftime("%H:%M:%S")
    start_perf = time.perf_counter()

    try:
        raw = sys.stdin.read().strip()
        if not raw:
            print(json.dumps({"success": False, "reason": "Empty JSON input"}))
            return

        params = json.loads(raw)

        periodfromdate = (params.get("fromdate") or "").strip()  # YYYY-MM-DD
        periodtodate   = (params.get("todate") or "").strip()    # YYYY-MM-DD
        comp           = int(params.get("companyid") or 0)
        out_dir        = params.get("out_dir") or os.getcwd()

        if not periodfromdate or not periodtodate or comp == 0:
            print(json.dumps({"success": False, "reason": "Missing fromdate/todate/companyid"}))
            return

        os.makedirs(out_dir, exist_ok=True)

        cn = get_connection()
        cur = cn.cursor(dictionary=True)

        try:
            # NOTE:
            # - We keep your query structure.
            # - Replace $comp and date strings with %s parameters.
            # - You are using ROW_NUMBER() in one UNION part => requires MySQL 8+.

            sql = """
            select * from (
                select 'Sales' `Voucher Type Name`,DATE_FORMAT(ih.invoice_date,'%d-%b-%Y') `Voucher Date` ,  ih.invoice_no_string `Voucher Number`
                ,ttlf1.tally_name `Ledger Name`,(ih.invoice_amount-ih.round_off)  `Ledger Amount`,'Dr' `Ledger Amount Dr/Cr` ,
                concat(' Ch.No: ',ih.challan_no) `Despatch Doc no.`,concat('Truck No: ',ih.vehicle_no) `Despath though`,
                ih.shipping_address   `Destination`,smh.mr_print_no  `other references`,' ' `Item Name`,' ' `Godown`,
                '' `Batch/Lotno.`,'' `Billed Quantity`,'' `Item Rate`,'' `Item Rate per`,
                '' `Item Amount`,'' `Description ?`,'Item Invoice' `Change Mode`,(ih.invoice_amount-ih.round_off)  `Bill Amount`,
                'Dr' `Bill Amount - Dr/Cr`,concat(ih.invoice_no_string,' / ',smh.mr_print_no) `Bill Name`,'New Ref' `Bill Type of Ref`,
                180 `Due or credit days`,  concat('Being ',ili.tqty,' Kg. Raw Jute Sold To ',ttlf1.tally_name,' against Inv No: ' ,
                  ih.invoice_no_string ,'  Dt: ',ih.invoice_date ,' MR NO: ',smh.mr_print_no ,', Amount Rs. ',ih.invoice_amount ,' /-')
                 Narration,cm.address  `Address1`,cm.state  State1 ,'India'  Country1 ,cm.address  Address2,cm.state2   State2,'India' Country2,
                 ih.invoice_date,ih.invoice_no_string,0 invoice_line_item_id,'hd' rem,0 slno
                from invoice_hdr ih
                left join customer_master cm  on cm.id =ih.customer_id
                left join EMPMILL12.tbl_tally_link_file ttlf1 on ih.company_id =ttlf1.company_id and ttlf1.link_for ='M'  and ttlf1.vow_name = cm.name
                left join scm_mr_hdr smh on ih.mr_id=smh.jute_receive_no
                left join (select ili.invoice_id,sum(ili.quantity*100) tqty from invoice_line_items ili where ili.is_active=1  group by ili.invoice_id  )
                ili on ih.invoice_id =ili.invoice_id
                where ih.status not in (4,6) and ih.is_active =1 and  ih.company_id = %s and ih.invoice_date between %s and %s

                UNION ALL

                select '' `Voucher Type Name`,'' `Voucher Date` ,  '' `Voucher Number`
                ,'Sale of Raw Jute' `Ledger Name`,(ih.invoice_amount-ih.round_off)  `Ledger Amount`,'Cr' `Ledger Amount Dr/Cr` ,
                '' `Despatch Doc no.`,'' `Despath though`,''  `Destination`,''  `other references`,
                ttlf2.tally_name  `Item Name`,ttlf.tally_name  `Godown`,
                smh.mr_print_no `Batch/Lotno.`,ili.quantity*100  `Billed Quantity`,ili.rate/100  `Item Rate`,'kg' `Item Rate per`,
                ili.amount_without_tax  `Item Amount`,CONCAT(
                  'Raw Jute: ',
                  CASE
                    WHEN ili.sales_bale > 0 THEN CONCAT(ili.sales_bale, ' Bales')
                    ELSE 'loose'
                  END
                )  `Description ?`,'' `Change Mode`,'' `Bill Amount`,
                '' `Bill Amount - Dr/Cr`,'' `Bill Name`,'' `Bill Type of Ref`,
                '' `Due or credit days`, ''
                 Narration,''  `Address1`,''  State1 ,''  Country1 ,''  Address2,''   State2,'' Country2
                , ih.invoice_date,ih.invoice_no_string,ili.invoice_line_item_id, 'ln' rem,
                  ROW_NUMBER() OVER (PARTITION BY ili.invoice_id ORDER BY ili.invoice_line_item_id) AS slno
                 from invoice_hdr ih
                left join invoice_line_items ili on ih.invoice_id =ili.invoice_id
                left join jute_quality_price_master jqpm on jqpm.id=ili.quality_id
                left join EMPMILL12.tbl_tally_link_file ttlf2 on ih.company_id =ttlf2.company_id and ttlf2.link_for ='Q'  and ttlf2.vowid  = jqpm.id
                left join EMPMILL12.tbl_tally_link_file ttlf on ih.company_id =ttlf.company_id and ttlf.link_for ='G'
                left join scm_mr_hdr smh on ih.mr_id=smh.jute_receive_no
                where ih.status not in (4,6) and ih.is_active =1 and  ili.is_active=1 and ih.company_id = %s and ih.invoice_date between %s and %s

                union all

                select '' `Voucher Type Name`,'' `Voucher Date` , '' `Voucher Number`
                ,'Claim on Gross Sales'  `Ledger Name`,0-tclm `Ledger Amount`,'' `Ledger Amount Dr/Cr` ,
                '' `Despatch Doc no.`,'' `Despath though`,
                ''   `Destination`,''  `other references`,' ' `Item Name`,' ' `Godown`,
                '' `Batch/Lotno.`,'' `Billed Quantity`,'' `Item Rate`,'' `Item Rate per`,
                '' `Item Amount`,'' `Description ?`,'' `Change Mode`,'' `Bill Amount`,
                '' `Bill Amount - Dr/Cr`,'' `Bill Name`,'' `Bill Type of Ref`,
                '' `Due or credit days`, ''
                 Narration,''  `Address1`,''  State1 ,''  Country1 ,''  Address2,''   State2,'' Country2,
                 ih.invoice_date,ih.invoice_no_string,0 invoice_line_item_id,'oth' rem,0 slno
                from invoice_hdr ih
                left join customer_master cm  on cm.id =ih.customer_id
                left join EMPMILL12.tbl_tally_link_file ttlf1 on ih.company_id =ttlf1.company_id and ttlf1.link_for ='M'  and ttlf1.vow_name = cm.name
                left join scm_mr_hdr smh on ih.mr_id=smh.jute_receive_no
                left join (select ili.invoice_id,sum(ili.claim_amount_dtl ) tclm from invoice_line_items ili where ili.is_active=1 group by ili.invoice_id  )
                ili on ih.invoice_id =ili.invoice_id
                where ih.status not in (4,6) and ih.is_active =1 and ih.company_id = %s and ih.invoice_date between %s and %s and ili.tclm>0
            ) g
            order by invoice_date,invoice_no_string,rem,invoice_line_item_id
            """

            qparams = [
                comp, periodfromdate, periodtodate,
                comp, periodfromdate, periodtodate,
                comp, periodfromdate, periodtodate
            ]

            cur.execute(sql, qparams)
            rows = cur.fetchall()

            # ---- Excel ----
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales"

            MAX_COLS = 31

            if rows:
                headers = list(rows[0].keys())[:MAX_COLS]
                ws.append(headers)

                for r in rows:
                    row_vals = [r.get(h) for h in headers]

                    slno = r.get("slno")  # must match your query column name exactly

                    if slno is not None and int(slno) > 1:
                        # blank columns 4,5,6 (1-based) => indexes 3,4,5
                        for idx in (3, 4, 5):
                            if idx < len(row_vals):
                                row_vals[idx] = ""

                    ws.append(row_vals)
            else:
                ws.append(["No data"])


            autosize(ws)

            fname = f"Tally_Sales_{comp}_{periodfromdate}_to_{periodtodate}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(out_dir, fname)
            wb.save(file_path)

            elapsed = round(time.perf_counter() - start_perf, 3)
            print(json.dumps({
                "success": True,
                "file_path": file_path,
                "download_name": fname,
                "rows": len(rows),
                "start_time": start_clock,
                "end_time": datetime.now().strftime("%H:%M:%S"),
                "execution_time_sec": elapsed
            }))

        finally:
            try: cur.close()
            except: pass
            try: cn.close()
            except: pass

    except Exception as e:
        traceback.print_exc(file=sys.stderr)
        elapsed = round(time.perf_counter() - start_perf, 3)
        print(json.dumps({
            "success": False,
            "reason": str(e),
            "start_time": start_clock,
            "end_time": datetime.now().strftime("%H:%M:%S"),
            "execution_time_sec": elapsed
        }))


if __name__ == "__main__":
    main()
